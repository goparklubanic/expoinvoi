import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


# Open file resources/IDD250510 clearance.xlsx
wb = load_workbook('resources/IDD250510 clearance.xlsx')
ws = wb.active

# Evaluate Column A to find what row contains word Total
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        # If cell value is None then put 20 into the cell
        if cell.value is None:
            cell.value = 20
        if cell.value is not None and str(cell.value).lower() == 'total':
            total_row = cell.row
            break
    else:
        continue
    break
print(f"Total found at row number: {total_row}");
# Save the file
wb.save('resources/IDD250510 clearance.xlsx')