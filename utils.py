import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
# Pandas options to show full data frame
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)  # Avoid line wrapping
pd.set_option('display.max_colwidth', None)

# Read File Names
def list_excel_files(folder_path='resources'):
    """Returns a list of Excel file names in the given folder."""
    excel_files = [
        f for f in os.listdir(folder_path)
        if f.endswith(('.xlsx', '.xls')) and os.path.isfile(os.path.join(folder_path, f))
    ]
    return excel_files

# Extracting Data
def process_clearance_file(file_name, folder_path='resources'):
    file_path = os.path.join(folder_path, file_name)
    header_row = 15
    data_start_row = header_row + 1
    skip_keywords = ["PO#", "SUBTOTAL", "Mounting", "Buyer Dia","All unpaid balance will be charged 1.5% per month."]

    all_data = []
    current_label = None

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Get header from row 15
        headers = [str(cell.value).strip() if cell.value is not None else f"Unnamed_{i}"
                   for i, cell in enumerate(ws[header_row])]

        # Process data rows from row 16 onward
        for row in ws.iter_rows(min_row=data_start_row):
            first_cell = row[0].value

            # Skip row if first column has one of the keywords
            if first_cell:
                first_cell_str = str(first_cell).strip()
                if any(keyword.lower() in first_cell_str.lower() for keyword in skip_keywords):
                    continue
                # if first_cell_str == empty then continue
                if first_cell_str == "":
                    continue
                # If it's a small number, it's a sub-row (keep label)
                if first_cell_str.isdigit() and len(first_cell_str) <= 2:
                    pass
                else:
                    current_label = first_cell_str

            # Read row values
            row_values = [cell.value for cell in row]
            row_dict = dict(zip(headers, row_values))
            row_dict["Label"] = current_label
            row_dict["File"] = file_name
            all_data.append(row_dict)

    except Exception as e:
        print(f"Error processing {file_name}: {e}")
        return pd.DataFrame()

    return pd.DataFrame(all_data)

def clean_clearance_dataframe(df):
    """Clean and normalize clearance DataFrame with specific logic."""
    keep_columns = ["PO#", "Item No.", "Metal", "Q'ty", "Total W't", "maklon", "total", "File"]

    # Match columns (case-insensitive, flexible naming)
    col_map = {}
    for col in df.columns:
        for target in keep_columns:
            if col.strip().lower() == target.strip().lower():
                col_map[target] = col
                break

    # Keep only matched columns
    df = df[[col_map[col] for col in keep_columns if col in col_map]]
    df = df.rename(columns={v: k for k, v in col_map.items()})  # Rename to standard

    # Process PO# column
    cleaned_rows = []
    temp_po = None
    for _, row in df.iterrows():
        po_val = row["PO#"]
        if po_val is None:
            continue

        po_str = str(po_val).strip()

        # Stop if we hit "TOTAL"
        if "total" in po_str.lower():
            break

        # Update or reuse PO# value
        if len(po_str) > 3:
            temp_po = po_str
        elif len(po_str) <= 3 and temp_po:
            row["PO#"] = temp_po

        cleaned_rows.append(row)
    
    final_df = pd.DataFrame(cleaned_rows)

    # Reorder columns to make "File" first
    columns_ordered = ["File"] + [col for col in keep_columns if col != "File" and col in final_df.columns]
    final_df = final_df[columns_ordered]

    return final_df

def save_dataframe_to_excel(df, output_path='result/Extracted.xlsx'):
    """Append DataFrame rows to an Excel file without writing headers."""
    output_dir = os.path.dirname(output_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if os.path.exists(output_path):
        # Load existing workbook and get active sheet
        wb = load_workbook(output_path)
        ws = wb.active
        # Append rows without headers
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        wb.save(output_path)
    else:
        # First time write, include headers
        df.to_excel(output_path, index=False)
    
    print(f"✅ Appended data to {output_path}")
