import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import sys

def createDataframe(file_name, folder_path='resources'):
    file_path = os.path.join(folder_path, file_name)
    header_row = 15
    last_row = fillempty(file_name)
    data_start_row = header_row + 1
    # print(last_row)
    # sys.exit();
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Find last non-empty cell in header row
    # last_col_idx = max((cell.column for cell in ws[header_row] if cell.value is not None), default=1)
    # last_col_letter = get_column_letter(last_col_idx)


    # Get header from row 15
    start_col = 1
    end_col = ws.max_column
    headers = [cell.value for cell in ws[header_row][start_col-1:end_col]]
    # print(headers)
    # sys.exit();

    # print the length of headers
    header_length = len(headers)
    last_col_letter = get_column_letter(header_length)
    # print(last_col_letter)
    # sys.exit();
    df = pd.read_excel(
        file_path, 
        sheet_name="PI",
        skiprows=header_row - 1,
        nrows = last_row - data_start_row,
        usecols=f"A:{last_col_letter}",
        engine='openpyxl'
        )
    
    df.columns = headers

    # df = df[["PO#", "Item No.", "Metal", "Q'ty", "Total w't", "maklon", "total"]]
    df = df[["PO#", "Item No.", "Metal", "Q'ty", "Total w't", "manufacturing","non us dia","total"]]
    return df

def stripRows(df):
    skip_keywords = ["PO#", "SUBTOTAL", "Mounting", "Buyer Dia"]
    all_data = []
    current_label = None

    # Loop Column PO#, skip row if contain skip_keywords
    for index, row in df.iterrows():
        first_cell = row["PO#"]
        if first_cell:
            first_cell_str = str(first_cell).strip()
            # If the length of the value more than 3, then store it into current_label.
            if len(first_cell_str) > 3:
                current_label = first_cell_str
            #  if the length of the value less than 3, replace with current_label
            elif len(first_cell_str) < 3:
                row["PO#"] = current_label

            if any(keyword.lower() in first_cell_str.lower() for keyword in skip_keywords):
                continue
            all_data.append(row)
    return pd.DataFrame(all_data)

def addBuyer(df, buyer_name):
    df["Buyer Name"] = buyer_name
    # re arrange the column order. put Buyer Name to the first column
    df = df[["Buyer Name", "PO#", "Item No.", "Metal", "Q'ty", "Total w't", "manufacturing", "non us dia","total"]]
    return df

def AppendToExcel(df, output_path='result/Extracted.xlsx'):
    output_dir = os.path.dirname(output_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if os.path.exists(output_path):
        # Load existing workbook and get active sheet
        wb = load_workbook(output_path)
        ws = wb.active
        # Append rows without headers
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        wb.save(output_path)
    else:
        # First time write, include headers
        df.to_excel(output_path, index=False)
    
    print(f"✅ Appended data to {output_path}")
    

def fillempty(file):
    wb = load_workbook(f'resources/{file}')
    ws = wb.active
    total_row = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            # If cell value is None then put 20 into the cell
            if cell.value is None:
                cell.value = 20
                # All unpaid balance will be charged 1.5% per month. 
            if cell.value is not None and str(cell.value).lower() == 'all unpaid balance will be charged 1.5% per month. ':
                total_row = cell.row
                break
        else:
            continue
        break
    trow = total_row - 3
    # print(f"Total found at row number: {trow}");
    # sys.exit();
    # Save the file
    wb.save(f'resources/{file}')
    return trow
